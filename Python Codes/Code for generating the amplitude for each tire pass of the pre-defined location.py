import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

# Ask the user for the divisor for odd columns
# Incase of using ABAQUS time period in step, you must make sure to select an interval-
# -so that it allign with your time period of the amplitude.
# For example: If you have 1000 in divisor and you need repitition of 1000, the time period will be 1 section in ABAQUS Step module
# A quick check it to look into the final value of the Amplitude that you generate with the python code that i have provided in-
# -this section.
divisor = float(input("Enter the number by which to divide the values in the odd columns: "))

# Ask the user for the number of repetitions
# If you need 1000 passes, you need to put 10000 number and vice versa
# This is very important in consideration of tyre passes
num_repetitions = int(input("Enter the number of times to repeat the range (e.g., 1000): "))

# Create output directories if they don't exist
excel_output_folder = "Generated Amplitude for ABAQUS"
graphs_output_folder = "Repititive Cycles for Each Relative Position"
os.makedirs(excel_output_folder, exist_ok=True)
os.makedirs(graphs_output_folder, exist_ok=True)

# Load the existing Excel file
# This is the backbone of whole repititive loading and it has a pre-defined pattern that i have incuded in this course
# You must have this Excel sheet in order to use it, and, get benifit from this existing python code
# Incase of more or less then10 passes, you actually need to change this Excel sheet as well
input_filename = 'Repititive Cycles.xlsx'
wb = load_workbook(input_filename)

# Select the active sheet (assuming the range is in the first sheet)
sheet = wb.active

# Create a new workbook for the output
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = 'Cycles'

# Copy the header row (first row)
for col in range(1, 21):  # Columns A to T
    new_ws.cell(row=1, column=col).value = sheet.cell(row=1, column=col).value

# Define the range to copy (A2:T11)
# You have to adjust the Excel sheet, here if you have more then or less then 10 tyre contact passes, as in this current project it has 10 passes only
# To do so, you need to just make sure to change the location of the columns only
start_row, end_row = 2, 11
start_col, end_col = 1, 20  # A to T

# Copy the range and repeat it based on user input
for repeat in range(num_repetitions):
    for row in range(start_row, end_row + 1):
        # Calculate the new row number in the new sheet
        new_row = (repeat * (end_row - start_row + 1)) + row

        for col in range(start_col, end_col + 1):
            if col % 2 != 0:  # Check if the column is odd (A=1, C=3, etc.)
                # Calculate and set the incrementing value divided by the divisor
                increment_value = (0.1 * (new_row - 1)) / divisor  # 0.1, 0.2, 0.3, etc., divided by the user input
                new_ws.cell(row=new_row, column=col).value = increment_value
            else:
                # Get the cell value from the original sheet for even columns
                original_value = sheet.cell(row=row, column=col).value
                new_ws.cell(row=new_row, column=col).value = original_value

# Save the new workbook in the "Generated Amplitude for ABAQUS" folder
excel_output_filename = os.path.join(excel_output_folder, 'Required Cycles for ABAQUS.xlsx')
new_wb.save(excel_output_filename)

# Generate a graph for each pair of columns (odd and even)
tyre_pass_number = 1  # Counter for naming graphs
for col in range(1, end_col, 2):  # Iterate over odd columns A, C, E, etc.
    odd_col = col
    even_col = col + 1

    # Extract data for the columns
    odd_data = [new_ws.cell(row=r, column=odd_col).value for r in range(2, (num_repetitions * (end_row - start_row + 1)) + 2)]
    even_data = [new_ws.cell(row=r, column=even_col).value for r in range(2, (num_repetitions * (end_row - start_row + 1)) + 2)]

    # Plot the data
    plt.figure(figsize=(8, 6))

    # Use a thicker line to make the curve more visible
    plt.plot(odd_data, even_data, linestyle='-', color='b', linewidth=1.5)  # Thicker line

    # Assign the custom title and labels
    plt.title(f"Tyre Pass {tyre_pass_number}")
    plt.xlabel(f"Column {chr(odd_col + 64)} (Divided by {divisor})")
    plt.ylabel(f"Column {chr(even_col + 64)}")
    plt.grid(True)

    # Save the graph with a descriptive name
    graph_filename = os.path.join(graphs_output_folder, f"Tyre Pass {tyre_pass_number}.png")
    plt.savefig(graph_filename, dpi=300)  # Save with higher resolution
    plt.close()  # Close the figure to free up memory

    # Increment the tyre pass number for the next graph
    tyre_pass_number += 1

# Print completion messages
print(f"Excel file saved as {excel_output_filename}")
print(f"Graphs saved in {graphs_output_folder}")
